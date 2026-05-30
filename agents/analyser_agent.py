import json
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import ollama
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    Image, HRFlowable, PageBreak, KeepTogether
)

# ── Brand colours ──────────────────────────────────────────────────────────────
C_PURPLE       = "#4B3FA0"
C_PURPLE_LIGHT = "#EEEDFE"
C_PURPLE_MID   = "#7F77DD"
C_GREEN        = "#1A6B3A"
C_GREEN_LIGHT  = "#E6F5EC"
C_AMBER        = "#7A4500"
C_AMBER_LIGHT  = "#FFF3CD"
C_RED          = "#8B1A1A"
C_RED_LIGHT    = "#FDEAEA"
C_GRAY         = "#4A4A4A"
C_GRAY_LIGHT   = "#F5F5F5"
C_BORDER       = "#DDDAE8"

CHART_PALETTE = [
    "#4B3FA0","#7F77DD","#1A6B3A","#7A4500",
    "#185FA5","#993C1D","#0F6E56","#6B1A6B",
    "#2D6A8F","#5A5A00"
]

PAGE_W, PAGE_H = A4
MARGIN    = 1.8 * cm
CONTENT_W = PAGE_W - 2 * MARGIN
OLLAMA_MODEL = "llama3.2"


# ── Helpers ────────────────────────────────────────────────────────────────────

def _safe_float(val) -> float:
    try:
        return float(str(val).replace(",", "").strip())
    except Exception:
        return 0.0

def _safe_str(val) -> str:
    v = str(val).strip() if val else ""
    return "" if v in ("None", "nan", "") else v

def _pkr(amount: float) -> str:
    if amount >= 1_000_000:
        return f"PKR {amount/1_000_000:.2f}M"
    if amount >= 1_000:
        return f"PKR {amount/1_000:.0f}K"
    return f"PKR {amount:,.0f}"

def _save_chart(fig, path: Path):
    fig.savefig(str(path), dpi=150, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close(fig)


# ══════════════════════════════════════════════════════════════════════════════
# AI INSIGHTS — Ollama
# ══════════════════════════════════════════════════════════════════════════════

def generate_ai_insights(entries: list[dict], kpis: dict) -> dict:
    """
    Call Ollama once with full context and get back four AI-written sections.
    Returns dict with keys: summary, anomalies, trends, recommendations.
    Falls back to placeholder text if Ollama is unavailable.
    """
    print("  → AI: Generating insights with Ollama...")

    flagged = [
        e for e in entries
        if "warn" in _safe_str(e.get("Validation Status")).lower()
        or "error" in _safe_str(e.get("Validation Status")).lower()
    ]

    context = {
        "total_requests":     kpis["total_requests"],
        "total_amount_pkr":   round(kpis["total_amount"]),
        "average_amount_pkr": round(kpis["avg_amount"]),
        "max_amount_pkr":     round(kpis["max_amount"]),
        "pass_rate_pct":      kpis["pass_rate"],
        "flagged_count":      len(flagged),
        "top_department":     kpis["top_dept"],
        "most_common_reason": kpis["most_common_reason"],
        "daily_counts":       kpis["daily_counts"],
        "department_counts":  kpis["dept_counts"],
        "reason_breakdown":   kpis["reason_counts"],
        "flagged_entries": [
            {
                "employee": _safe_str(e.get("Employee Index")),
                "name":     _safe_str(e.get("Name")),
                "amount":   _safe_float(e.get("Amount Requested (PKR)")),
                "reason":   _safe_str(e.get("Reason")),
                "status":   _safe_str(e.get("Validation Status")),
            }
            for e in flagged
        ],
    }

    prompt = f"""You are a financial analyst writing a section of a Provident Fund report for a Pakistani company's HR/Finance department.

Here is the data for this period:
{json.dumps(context, indent=2)}

Write exactly four sections. Return ONLY a JSON object with these four keys — no markdown, no preamble:

{{
  "executive_summary": "3-4 sentences summarising the period. Mention total requests, total amount, top reason, and overall validation health. Write in professional business English.",
  
  "anomaly_commentary": "2-4 sentences specifically about the {len(flagged)} flagged entries. Name the issues found (e.g. balance exceeded, high value). If no flagged entries, write 'All entries passed validation with no anomalies detected.'",
  
  "trend_insights": "2-3 sentences about patterns in the data — which days are busiest, which department requests most, whether amounts are concentrated or spread out. Be specific using the numbers.",
  
  "recommendations": "3-4 bullet points (use - as bullet) of actionable recommendations for management based on the data. Examples: policy changes, process improvements, follow-up actions needed."
}}"""

    try:
        response = ollama.chat(
            model=OLLAMA_MODEL,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = response["message"]["content"].strip()
        # Extract JSON even if model adds surrounding text
        start = raw.find("{")
        end   = raw.rfind("}") + 1
        if start == -1 or end == 0:
            raise ValueError("No JSON object found in response")
        result = json.loads(raw[start:end])

        # Validate all four keys exist
        required = ["executive_summary", "anomaly_commentary",
                    "trend_insights", "recommendations"]
        for key in required:
            if key not in result:
                result[key] = _fallback_insights(kpis, flagged)[key]

        print("  → AI: Insights generated successfully")
        return result

    except Exception as e:
        print(f"  → AI: Ollama unavailable ({e}), using rule-based insights")
        return _fallback_insights(kpis, flagged)


def _fallback_insights(kpis: dict, flagged: list) -> dict:
    """Rule-based fallback if Ollama is not running."""
    flag_text = (
        f"{len(flagged)} entries were flagged requiring review — "
        f"including issues such as balance exceedance and high-value withdrawals."
        if flagged else
        "All entries passed validation with no anomalies detected."
    )
    return {
        "executive_summary": (
            f"This period recorded {kpis['total_requests']} PF withdrawal requests "
            f"totalling {_pkr(kpis['total_amount'])}. "
            f"The average withdrawal was {_pkr(kpis['avg_amount'])}. "
            f"The validation pass rate stood at {kpis['pass_rate']}%."
        ),
        "anomaly_commentary": flag_text,
        "trend_insights": (
            f"The {kpis['top_dept']} department submitted the highest number of requests. "
            f"The most frequently cited reason was '{kpis['most_common_reason']}'. "
            f"Withdrawal amounts ranged from {_pkr(kpis['min_amount'])} "
            f"to {_pkr(kpis['max_amount'])}."
        ),
        "recommendations": (
            "- Review all flagged entries before period close.\n"
            "- Remind employees to check their PF balance before submitting requests.\n"
            "- Consider adding a balance check to the submission form.\n"
            "- Share this report with department heads for their review."
        ),
    }


# ══════════════════════════════════════════════════════════════════════════════
# KPI CALCULATION
# ══════════════════════════════════════════════════════════════════════════════

def calculate_kpis(entries: list[dict]) -> dict:
    if not entries:
        return {}

    amounts_pos = [a for a in
                   [_safe_float(e.get("Amount Requested (PKR)")) for e in entries]
                   if a > 0]
    reasons  = [_safe_str(e.get("Reason")) for e in entries if _safe_str(e.get("Reason"))]
    statuses = [_safe_str(e.get("Validation Status")) for e in entries]

    daily: dict[str, int] = defaultdict(int)
    for e in entries:
        daily[_safe_str(e.get("_sheet"))] += 1

    reason_counts: dict[str, int] = defaultdict(int)
    for r in reasons:
        reason_counts[" ".join(r.split()[:4])] += 1

    dept_amounts: dict[str, float] = defaultdict(float)
    dept_counts:  dict[str, int]   = defaultdict(int)
    for e in entries:
        dept = _safe_str(e.get("Department")) or "Unknown"
        dept_amounts[dept] += _safe_float(e.get("Amount Requested (PKR)"))
        dept_counts[dept]  += 1

    ok_count   = sum(1 for s in statuses if "ok" in s.lower() or
                     ("valid" in s.lower() and "invalid" not in s.lower()))
    warn_count = sum(1 for s in statuses if "warn" in s.lower())
    err_count  = sum(1 for s in statuses if "error" in s.lower())

    return {
        "total_requests":     len(entries),
        "total_amount":       sum(amounts_pos),
        "avg_amount":         sum(amounts_pos) / len(amounts_pos) if amounts_pos else 0,
        "max_amount":         max(amounts_pos) if amounts_pos else 0,
        "min_amount":         min(amounts_pos) if amounts_pos else 0,
        "daily_counts":       dict(sorted(daily.items())),
        "reason_counts":      dict(sorted(reason_counts.items(), key=lambda x: -x[1])[:7]),
        "dept_amounts":       dict(sorted(dept_amounts.items(), key=lambda x: -x[1])),
        "dept_counts":        dict(sorted(dept_counts.items(), key=lambda x: -x[1])),
        "ok_count":           ok_count,
        "warn_count":         warn_count,
        "err_count":          err_count,
        "pass_rate":          round(ok_count / len(entries) * 100, 1) if entries else 0,
        "most_common_reason": max(reason_counts, key=reason_counts.get) if reason_counts else "N/A",
        "top_dept":           max(dept_counts,   key=dept_counts.get)   if dept_counts   else "N/A",
    }


# ══════════════════════════════════════════════════════════════════════════════
# CHARTS
# ══════════════════════════════════════════════════════════════════════════════

def _apply_spine_style(ax):
    for spine in ax.spines.values():
        spine.set_color(C_BORDER)
        spine.set_linewidth(0.8)
    ax.set_facecolor("#FAFAFA")
    ax.grid(axis="y", linestyle="--", alpha=0.45, color=C_BORDER, zorder=0)
    ax.tick_params(colors=C_GRAY, labelsize=8)


def chart_daily_requests(kpis: dict, out_path: Path):
    daily = kpis.get("daily_counts", {})
    if not daily:
        return
    dates  = list(daily.keys())
    counts = list(daily.values())
    fig, ax = plt.subplots(figsize=(11, 3.8))
    bars = ax.bar(range(len(dates)), counts, color=C_PURPLE, width=0.55, zorder=3)
    ax.set_xticks(range(len(dates)))
    ax.set_xticklabels(dates, rotation=35, ha="right", fontsize=8, color=C_GRAY)
    ax.set_ylabel("Requests", fontsize=9, color=C_GRAY, labelpad=8)
    ax.set_title("Daily PF Requests", fontsize=12, fontweight="bold",
                 color=C_PURPLE, pad=14, loc="left")
    ax.yaxis.set_major_locator(plt.MaxNLocator(integer=True))
    _apply_spine_style(ax)
    for bar, val in zip(bars, counts):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.06,
                str(val), ha="center", va="bottom",
                fontsize=8, fontweight="bold", color=C_PURPLE)
    fig.tight_layout(pad=1.5)
    _save_chart(fig, out_path)


def chart_reasons_pie(kpis: dict, out_path: Path):
    reasons = kpis.get("reason_counts", {})
    if not reasons:
        return
    labels = list(reasons.keys())
    sizes  = list(reasons.values())
    clrs   = CHART_PALETTE[:len(labels)]
    fig, ax = plt.subplots(figsize=(7, 4.2))
    wedges, _, autotexts = ax.pie(
        sizes, colors=clrs, autopct="%1.0f%%", startangle=140,
        pctdistance=0.78,
        wedgeprops={"linewidth": 1.5, "edgecolor": "white"},
    )
    for at in autotexts:
        at.set_fontsize(8)
        at.set_color("white")
        at.set_fontweight("bold")
    ax.legend(wedges, labels, loc="lower center",
              bbox_to_anchor=(0.5, -0.22), ncol=2,
              fontsize=7.5, frameon=False, labelcolor=C_GRAY)
    ax.set_title("Withdrawal Reasons", fontsize=12, fontweight="bold",
                 color=C_PURPLE, pad=12)
    fig.tight_layout(pad=1.5)
    _save_chart(fig, out_path)


def chart_department(kpis: dict, out_path: Path):
    dept_c = kpis.get("dept_counts", {})
    dept_a = kpis.get("dept_amounts", {})
    if not dept_c:
        return
    depts   = list(dept_c.keys())
    counts  = [dept_c[d] for d in depts]
    amounts = [dept_a.get(d, 0) / 1000 for d in depts]
    y_pos   = range(len(depts))
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(11, max(3.5, len(depts) * 0.65)))
    bars1 = ax1.barh(list(y_pos), counts, color=C_PURPLE, height=0.5, zorder=3)
    ax1.set_yticks(list(y_pos))
    ax1.set_yticklabels(depts, fontsize=9, color=C_GRAY)
    ax1.set_xlabel("Requests", fontsize=8, color=C_GRAY)
    ax1.set_title("Requests by Department", fontsize=11,
                  fontweight="bold", color=C_PURPLE, pad=10, loc="left")
    ax1.xaxis.set_major_locator(plt.MaxNLocator(integer=True))
    _apply_spine_style(ax1)
    ax1.grid(axis="x", linestyle="--", alpha=0.45, color=C_BORDER, zorder=0)
    ax1.grid(axis="y", visible=False)
    for bar, val in zip(bars1, counts):
        ax1.text(bar.get_width() + 0.05, bar.get_y() + bar.get_height() / 2,
                 str(val), va="center", fontsize=8, fontweight="bold", color=C_PURPLE)
    bars2 = ax2.barh(list(y_pos), amounts, color=C_PURPLE_MID, height=0.5, zorder=3)
    ax2.set_yticks(list(y_pos))
    ax2.set_yticklabels([], fontsize=9)
    ax2.set_xlabel("Amount (PKR '000)", fontsize=8, color=C_GRAY)
    ax2.set_title("Amount by Department", fontsize=11,
                  fontweight="bold", color=C_PURPLE, pad=10, loc="left")
    _apply_spine_style(ax2)
    ax2.grid(axis="x", linestyle="--", alpha=0.45, color=C_BORDER, zorder=0)
    ax2.grid(axis="y", visible=False)
    for bar, val in zip(bars2, amounts):
        ax2.text(bar.get_width() + max(amounts) * 0.01,
                 bar.get_y() + bar.get_height() / 2,
                 f"{val:,.0f}k", va="center", fontsize=8, color=C_GRAY)
    fig.tight_layout(pad=1.5)
    _save_chart(fig, out_path)


def chart_validation(kpis: dict, out_path: Path):
    labels  = ["Valid", "Warnings", "Errors"]
    values  = [kpis["ok_count"], kpis["warn_count"], kpis["err_count"]]
    clrs    = [C_GREEN, C_AMBER, C_RED]
    total   = sum(values) or 1
    fig, ax = plt.subplots(figsize=(4.5, 3.8))
    bars = ax.bar(labels, values, color=clrs, width=0.42, zorder=3)
    ax.set_title("Validation Summary", fontsize=11,
                 fontweight="bold", color=C_PURPLE, pad=12, loc="left")
    ax.set_ylabel("Count", fontsize=9, color=C_GRAY)
    ax.yaxis.set_major_locator(plt.MaxNLocator(integer=True))
    _apply_spine_style(ax)
    for bar, val in zip(bars, values):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.05,
                f"{val}\n({val/total*100:.0f}%)",
                ha="center", va="bottom", fontsize=8,
                fontweight="bold", color=C_GRAY)
    fig.tight_layout(pad=1.5)
    _save_chart(fig, out_path)


# ══════════════════════════════════════════════════════════════════════════════
# PDF STYLES
# ══════════════════════════════════════════════════════════════════════════════

def _styles():
    return {
        "report_title": ParagraphStyle("RT",
            fontName="Helvetica-Bold", fontSize=24,
            textColor=colors.HexColor(C_PURPLE),
            alignment=TA_LEFT, leading=28, spaceAfter=2),

        "report_subtitle": ParagraphStyle("RS",
            fontName="Helvetica", fontSize=11,
            textColor=colors.HexColor(C_GRAY),
            alignment=TA_LEFT, leading=15),

        "meta_right": ParagraphStyle("MR",
            fontName="Helvetica", fontSize=9,
            textColor=colors.HexColor("#999999"),
            alignment=TA_RIGHT, leading=13),

        "section": ParagraphStyle("SEC",
            fontName="Helvetica-Bold", fontSize=13,
            textColor=colors.HexColor(C_PURPLE),
            spaceBefore=14, spaceAfter=6, leading=16),

        "ai_section": ParagraphStyle("AIS",
            fontName="Helvetica-Bold", fontSize=11,
            textColor=colors.HexColor(C_PURPLE),
            spaceBefore=10, spaceAfter=4, leading=14),

        "ai_body": ParagraphStyle("AIB",
            fontName="Helvetica", fontSize=9.5,
            textColor=colors.HexColor(C_GRAY),
            spaceAfter=4, leading=15),

        "ai_bullet": ParagraphStyle("AIBUL",
            fontName="Helvetica", fontSize=9.5,
            textColor=colors.HexColor(C_GRAY),
            spaceAfter=3, leading=15,
            leftIndent=12, bulletIndent=0),

        "body": ParagraphStyle("BD",
            fontName="Helvetica", fontSize=9,
            textColor=colors.HexColor(C_GRAY),
            spaceAfter=4, leading=13),

        "kpi_value": ParagraphStyle("KV",
            fontName="Helvetica-Bold", fontSize=18,
            textColor=colors.HexColor(C_PURPLE),
            alignment=TA_CENTER, leading=22, spaceAfter=0),

        "kpi_label": ParagraphStyle("KL",
            fontName="Helvetica", fontSize=8,
            textColor=colors.HexColor(C_GRAY),
            alignment=TA_CENTER, leading=11),

        "table_hdr": ParagraphStyle("TH",
            fontName="Helvetica-Bold", fontSize=8,
            textColor=colors.white, alignment=TA_CENTER),

        "table_cell": ParagraphStyle("TC",
            fontName="Helvetica", fontSize=8,
            textColor=colors.HexColor(C_GRAY)),

        "table_amt": ParagraphStyle("TA",
            fontName="Helvetica", fontSize=8,
            textColor=colors.HexColor(C_GRAY), alignment=TA_RIGHT),

        "no_issues": ParagraphStyle("NI",
            fontName="Helvetica-Bold", fontSize=11,
            textColor=colors.HexColor(C_GREEN),
            alignment=TA_CENTER, spaceBefore=20),

        "footer": ParagraphStyle("FT",
            fontName="Helvetica", fontSize=7.5,
            textColor=colors.HexColor("#AAAAAA"),
            alignment=TA_CENTER),

        "ai_badge": ParagraphStyle("AB",
            fontName="Helvetica-Bold", fontSize=7.5,
            textColor=colors.HexColor(C_PURPLE_MID),
            alignment=TA_RIGHT),
    }


# ══════════════════════════════════════════════════════════════════════════════
# PDF BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def _header_band(text: str, S: dict, bg: str = C_PURPLE_LIGHT) -> Table:
    t = Table([[Paragraph(text, S["report_title"])]],
              colWidths=[CONTENT_W], rowHeights=[1.3*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), colors.HexColor(bg)),
        ("LEFTPADDING",   (0,0), (-1,-1), 14),
        ("TOPPADDING",    (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("ROUNDEDCORNERS", [6]),
    ]))
    return t


def _kpi_card(value_str: str, label: str, bg: str, S: dict, width: float) -> Table:
    t = Table(
        [[Paragraph(value_str, S["kpi_value"])],
         [Paragraph(label,     S["kpi_label"])]],
        colWidths=[width],
        rowHeights=[1.05*cm, 0.5*cm],
    )
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), colors.HexColor(bg)),
        ("TOPPADDING",    (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("LEFTPADDING",   (0,0), (-1,-1), 4),
        ("RIGHTPADDING",  (0,0), (-1,-1), 4),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("BOX",           (0,0), (-1,-1), 0.5, colors.HexColor(C_BORDER)),
        ("ROUNDEDCORNERS", [5]),
    ]))
    return t


def _ai_box(title: str, content: str, S: dict,
            bg: str = "#F8F7FF", border: str = C_PURPLE_MID) -> Table:
    """Render an AI insight in a styled box."""
    # Split bullet points if present
    if "\n-" in content or content.startswith("-"):
        lines = [l.strip() for l in content.split("\n") if l.strip()]
        paragraphs = []
        for line in lines:
            if line.startswith("-"):
                paragraphs.append(
                    Paragraph("• " + line[1:].strip(), S["ai_bullet"])
                )
            else:
                paragraphs.append(Paragraph(line, S["ai_body"]))
        body_content = paragraphs
    else:
        body_content = [Paragraph(content, S["ai_body"])]

    header_cell = Table(
        [[Paragraph(title, S["ai_section"]),
          Paragraph("✦ AI Generated", S["ai_badge"])]],
        colWidths=[CONTENT_W * 0.75, CONTENT_W * 0.25 - 1.2*cm]
    )
    header_cell.setStyle(TableStyle([
        ("LEFTPADDING",  (0,0), (-1,-1), 0),
        ("RIGHTPADDING", (0,0), (-1,-1), 0),
        ("TOPPADDING",   (0,0), (-1,-1), 0),
        ("BOTTOMPADDING",(0,0), (-1,-1), 0),
        ("VALIGN",       (0,0), (-1,-1), "BOTTOM"),
    ]))

    inner_rows = [[header_cell]] + [[p] for p in body_content]
    inner_widths = [CONTENT_W - 1.2*cm]
    inner = Table(inner_rows, colWidths=inner_widths)
    inner.setStyle(TableStyle([
        ("LEFTPADDING",  (0,0), (-1,-1), 0),
        ("RIGHTPADDING", (0,0), (-1,-1), 0),
        ("TOPPADDING",   (0,0), (-1,-1), 0),
        ("BOTTOMPADDING",(0,0), (-1,-1), 2),
    ]))

    box = Table([[inner]], colWidths=[CONTENT_W])
    box.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), colors.HexColor(bg)),
        ("LEFTPADDING",   (0,0), (-1,-1), 14),
        ("RIGHTPADDING",  (0,0), (-1,-1), 14),
        ("TOPPADDING",    (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 12),
        ("BOX",           (0,0), (-1,-1), 1.2, colors.HexColor(border)),
        ("ROUNDEDCORNERS", [6]),
    ]))
    return box


def generate_pdf(entries: list[dict], kpis: dict, insights: dict,
                 output_path: Path, period_label: str, tmp_dir: Path) -> Path:

    tmp_dir.mkdir(parents=True, exist_ok=True)
    S = _styles()

    # Generate charts
    cp = {
        "daily":      tmp_dir / "c_daily.png",
        "reasons":    tmp_dir / "c_reasons.png",
        "dept":       tmp_dir / "c_dept.png",
        "validation": tmp_dir / "c_validation.png",
    }
    chart_daily_requests(kpis, cp["daily"])
    chart_reasons_pie(kpis, cp["reasons"])
    chart_department(kpis, cp["dept"])
    chart_validation(kpis, cp["validation"])

    doc = SimpleDocTemplate(
        str(output_path), pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=1.6*cm, bottomMargin=1.6*cm,
        title="Provident Fund Report",
        author="PF Automation Agent",
    )
    story = []
    card_w = (CONTENT_W - 3 * 0.25*cm) / 4

    # ═══════════════════════════════════════════════════════════════════════
    # PAGE 1 — Header + KPI Cards + AI Executive Summary + Quick Stats
    # ═══════════════════════════════════════════════════════════════════════

    story.append(_header_band("Provident Fund Report", S))
    story.append(Spacer(1, 0.2*cm))

    # Period + date on same line
    meta = Table(
        [[Paragraph(f"<b>Period:</b> {period_label}", S["report_subtitle"]),
          Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}",
                    S["meta_right"])]],
        colWidths=[CONTENT_W * 0.6, CONTENT_W * 0.4],
    )
    meta.setStyle(TableStyle([
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING",  (0,0), (-1,-1), 0),
        ("RIGHTPADDING", (0,0), (-1,-1), 0),
        ("TOPPADDING",   (0,0), (-1,-1), 0),
        ("BOTTOMPADDING",(0,0), (-1,-1), 0),
    ]))
    story.append(meta)
    story.append(Spacer(1, 0.15*cm))
    story.append(HRFlowable(width="100%", thickness=1.5,
                             color=colors.HexColor(C_PURPLE), spaceAfter=12))

    # ── KPI Row 1 ──
    story.append(Paragraph("Key Performance Indicators", S["section"]))
    story.append(Spacer(1, 0.1*cm))

    row1 = Table([[
        _kpi_card(_pkr(kpis["total_amount"]),  "Total Disbursed",       C_PURPLE_LIGHT, S, card_w),
        _kpi_card(str(kpis["total_requests"]), "Total Requests",         C_PURPLE_LIGHT, S, card_w),
        _kpi_card(_pkr(kpis["avg_amount"]),    "Avg. Per Request",       C_PURPLE_LIGHT, S, card_w),
        _kpi_card(f"{kpis['pass_rate']}%",     "Validation Pass Rate",   C_PURPLE_LIGHT, S, card_w),
    ]], colWidths=[card_w]*4)
    row1.setStyle(TableStyle([
        ("LEFTPADDING",  (0,0), (-1,-1), 3),
        ("RIGHTPADDING", (0,0), (-1,-1), 3),
        ("TOPPADDING",   (0,0), (-1,-1), 0),
        ("BOTTOMPADDING",(0,0), (-1,-1), 0),
        ("VALIGN",       (0,0), (-1,-1), "TOP"),
    ]))
    story.append(row1)
    story.append(Spacer(1, 0.25*cm))

    # ── KPI Row 2 ──
    row2 = Table([[
        _kpi_card(_pkr(kpis["max_amount"]), "Highest Withdrawal", C_RED_LIGHT,   S, card_w),
        _kpi_card(_pkr(kpis["min_amount"]), "Lowest Withdrawal",  C_GREEN_LIGHT, S, card_w),
        _kpi_card(str(kpis["warn_count"]),  "Warnings",           C_AMBER_LIGHT, S, card_w),
        _kpi_card(str(kpis["err_count"]),   "Errors",             C_RED_LIGHT,   S, card_w),
    ]], colWidths=[card_w]*4)
    row2.setStyle(TableStyle([
        ("LEFTPADDING",  (0,0), (-1,-1), 3),
        ("RIGHTPADDING", (0,0), (-1,-1), 3),
        ("TOPPADDING",   (0,0), (-1,-1), 0),
        ("BOTTOMPADDING",(0,0), (-1,-1), 0),
        ("VALIGN",       (0,0), (-1,-1), "TOP"),
    ]))
    story.append(row2)
    story.append(Spacer(1, 0.4*cm))
    story.append(HRFlowable(width="100%", thickness=0.5,
                             color=colors.HexColor(C_BORDER), spaceAfter=10))

    # ── AI Executive Summary ──
    story.append(_ai_box(
        "Executive Summary",
        insights["executive_summary"],
        S, bg="#F8F7FF", border=C_PURPLE_MID
    ))
    story.append(Spacer(1, 0.35*cm))

    # ── Quick Stats Table ──
    story.append(Paragraph("Summary Statistics", S["section"]))
    qs_data = [
        ["Metric", "Value"],
        ["Most Common Reason",  kpis["most_common_reason"]],
        ["Top Department",      kpis["top_dept"]],
        ["Total Requests",      str(kpis["total_requests"])],
        ["Total Amount",        f"PKR {kpis['total_amount']:,.0f}"],
        ["Average Withdrawal",  f"PKR {kpis['avg_amount']:,.0f}"],
        ["Highest Withdrawal",  f"PKR {kpis['max_amount']:,.0f}"],
        ["Lowest Withdrawal",   f"PKR {kpis['min_amount']:,.0f}"],
        ["Valid Entries",       f"{kpis['ok_count']} ({kpis['pass_rate']}%)"],
        ["Flagged Entries",     str(kpis['warn_count'] + kpis['err_count'])],
    ]
    qs = Table(qs_data, colWidths=[CONTENT_W*0.42, CONTENT_W*0.58])
    qs.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), colors.HexColor(C_PURPLE)),
        ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,-1), 9),
        ("FONTNAME",      (0,1), (0,-1), "Helvetica-Bold"),
        ("TEXTCOLOR",     (0,1), (0,-1), colors.HexColor(C_PURPLE)),
        ("TEXTCOLOR",     (1,1), (1,-1), colors.HexColor(C_GRAY)),
        ("ROWBACKGROUNDS",(0,1), (-1,-1),
         [colors.HexColor(C_PURPLE_LIGHT), colors.white]),
        ("TOPPADDING",    (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("LEFTPADDING",   (0,0), (-1,-1), 10),
        ("RIGHTPADDING",  (0,0), (-1,-1), 10),
        ("LINEBELOW",     (0,0), (-1,-2), 0.3, colors.HexColor(C_BORDER)),
        ("BOX",           (0,0), (-1,-1), 0.5, colors.HexColor(C_BORDER)),
    ]))
    story.append(qs)

    # ═══════════════════════════════════════════════════════════════════════
    # PAGE 2 — Charts + AI Insights
    # ═══════════════════════════════════════════════════════════════════════
    story.append(PageBreak())
    story.append(_header_band("Charts &amp; AI Insights", S))
    story.append(Spacer(1, 0.3*cm))

    # Daily chart
    if cp["daily"].exists():
        story.append(Paragraph("Daily Request Volume", S["section"]))
        story.append(Image(str(cp["daily"]), width=CONTENT_W, height=5.2*cm))
        story.append(Spacer(1, 0.35*cm))

    # AI Trend Insights
    story.append(_ai_box(
        "Trend Insights",
        insights["trend_insights"],
        S, bg="#F0FFF4", border=C_GREEN
    ))
    story.append(Spacer(1, 0.4*cm))
    story.append(HRFlowable(width="100%", thickness=0.5,
                             color=colors.HexColor(C_BORDER), spaceAfter=10))

    # Reasons + Validation side by side
    chart_cells = []
    chart_widths = []
    if cp["reasons"].exists():
        chart_cells.append(Image(str(cp["reasons"]), width=CONTENT_W*0.62, height=5.2*cm))
        chart_widths.append(CONTENT_W * 0.63)
    if cp["validation"].exists():
        chart_cells.append(Image(str(cp["validation"]), width=CONTENT_W*0.35, height=5.2*cm))
        chart_widths.append(CONTENT_W * 0.37)
    if chart_cells:
        ct = Table([chart_cells], colWidths=chart_widths)
        ct.setStyle(TableStyle([
            ("VALIGN",       (0,0), (-1,-1), "TOP"),
            ("LEFTPADDING",  (0,0), (-1,-1), 0),
            ("RIGHTPADDING", (0,0), (-1,-1), 0),
        ]))
        story.append(ct)
        story.append(Spacer(1, 0.4*cm))

    # Dept chart
    if cp["dept"].exists():
        story.append(HRFlowable(width="100%", thickness=0.5,
                                 color=colors.HexColor(C_BORDER), spaceAfter=10))
        story.append(Image(str(cp["dept"]), width=CONTENT_W, height=5.2*cm))

    # ═══════════════════════════════════════════════════════════════════════
    # PAGE 3 — Flagged Entries + AI Analysis + Recommendations
    # ═══════════════════════════════════════════════════════════════════════
    story.append(PageBreak())

    flagged = [
        e for e in entries
        if "warn" in _safe_str(e.get("Validation Status")).lower()
        or "error" in _safe_str(e.get("Validation Status")).lower()
    ]

    story.append(_header_band(
        f"Flagged Entries &amp; Recommendations",
        S,
        bg=C_RED_LIGHT if flagged else C_GREEN_LIGHT
    ))
    story.append(Spacer(1, 0.35*cm))

    # AI Anomaly Commentary
    story.append(_ai_box(
        "AI Anomaly Analysis",
        insights["anomaly_commentary"],
        S, bg="#FFF8F8", border=C_RED
    ))
    story.append(Spacer(1, 0.35*cm))

    # Flagged entries table
    if flagged:
        story.append(Paragraph(f"Flagged Entries — {len(flagged)} requiring attention",
                               S["section"]))
        col_w = [0.8*cm, 2.3*cm, 2.5*cm, 3.3*cm, 2.0*cm, 2.6*cm, 3.3*cm]
        flag_data = [[
            Paragraph(h, S["table_hdr"])
            for h in ["#", "Date", "Emp. Index", "Name", "Dept", "Amount", "Status"]
        ]]
        for e in flagged:
            status    = _safe_str(e.get("Validation Status"))
            is_err    = "error" in status.lower()
            flag_data.append([
                Paragraph(_safe_str(e.get("#")),              S["table_cell"]),
                Paragraph(_safe_str(e.get("_sheet")),         S["table_cell"]),
                Paragraph(_safe_str(e.get("Employee Index")), S["table_cell"]),
                Paragraph(_safe_str(e.get("Name")),           S["table_cell"]),
                Paragraph(_safe_str(e.get("Department")),     S["table_cell"]),
                Paragraph(
                    f"PKR {_safe_float(e.get('Amount Requested (PKR)')):,.0f}",
                    S["table_amt"]
                ),
                Paragraph(status, ParagraphStyle("SC",
                    parent=S["table_cell"],
                    textColor=colors.HexColor(C_RED if is_err else C_AMBER),
                    fontName="Helvetica-Bold")),
            ])
        ft = Table(flag_data, colWidths=col_w, repeatRows=1)
        ft.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), colors.HexColor(C_PURPLE)),
            ("FONTSIZE",      (0,0), (-1,-1), 8),
            ("ROWBACKGROUNDS",(0,1), (-1,-1),
             [colors.white, colors.HexColor(C_GRAY_LIGHT)]),
            ("TOPPADDING",    (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
            ("LEFTPADDING",   (0,0), (-1,-1), 6),
            ("RIGHTPADDING",  (0,0), (-1,-1), 6),
            ("GRID",          (0,0), (-1,-1), 0.3, colors.HexColor(C_BORDER)),
            ("BOX",           (0,0), (-1,-1), 0.8, colors.HexColor(C_PURPLE)),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ]))
        story.append(ft)
    else:
        story.append(Paragraph(
            "✓  All entries passed validation. No issues found.",
            S["no_issues"]
        ))

    story.append(Spacer(1, 0.45*cm))

    # AI Recommendations
    story.append(_ai_box(
        "Recommendations",
        insights["recommendations"],
        S, bg="#FFFDF0", border=C_AMBER
    ))

    # Footer
    story.append(Spacer(1, 1*cm))
    story.append(HRFlowable(width="100%", thickness=0.5,
                             color=colors.HexColor(C_BORDER), spaceAfter=6))
    story.append(Paragraph(
        f"Generated automatically by PF Automation Agent on "
        f"{datetime.now().strftime('%d %B %Y at %H:%M')}. "
        f"AI insights powered by Ollama ({OLLAMA_MODEL}). "
        f"For queries contact your HR/Finance department.",
        S["footer"]
    ))

    doc.build(story)
    return output_path


# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════

def load_all_data(data_dir: Path, filter_month: str = None) -> list[dict]:
    all_entries = []
    xlsx_files  = sorted(data_dir.glob("PF_Requests_*.xlsx"))
    if not xlsx_files:
        print(f"No Excel files found in {data_dir}")
        return []
    for xlsx_path in xlsx_files:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            if filter_month and not sheet_name.startswith(filter_month):
                continue
            ws   = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue
            headers = [str(h).strip() if h else "" for h in rows[0]]
            for row_idx, row in enumerate(rows[1:], start=2):
                entry = dict(zip(headers, row))
                entry["_sheet"] = sheet_name
                entry["_file"]  = xlsx_path.name
                all_entries.append(entry)
        wb.close()
    print(f"Loaded {len(all_entries)} entries from {len(xlsx_files)} Excel file(s)")
    return all_entries


# need openpyxl for load_all_data
import openpyxl


# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def run_analysis(
    filter_month=None,
    analyse_all=False
):
    data_dir = Path("data")
    output_dir = Path("reports")

    entries = load_all_data(
        data_dir=data_dir,
        filter_month=filter_month
    )

    period_label = filter_month if filter_month else "All Periods"

    output_dir.mkdir(parents=True, exist_ok=True)

    kpis = calculate_kpis(entries)
    insights = generate_ai_insights(entries, kpis)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    out_path = output_dir / f"PF_Report_{timestamp}.pdf"

    generate_pdf(
        entries,
        kpis,
        insights,
        out_path,
        period_label,
        output_dir / "tmp_charts"
    )

    return out_path
