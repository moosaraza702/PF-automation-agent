"""
Excel manager: daily sheet creation, file rotation, and entry writing.

Sheet structure per day:
  A: #  B: Timestamp  C: Employee Index  D: Name  E: Department
  F: Grade  G: PF Balance  H: Amount Requested  I: Reason
  J: Source  K: Confidence  L: Gmail Message ID  M: Validation Status
"""

from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from loguru import logger

from config.settings import EXCEL_OUTPUT_DIR, EXCEL_ROTATION_MONTHS

COLUMNS = [
    "#", "Timestamp", "Employee Index", "Name", "Department",
    "Grade", "PF Balance (PKR)", "Amount Requested (PKR)", "Reason",
    "Source", "Confidence", "Gmail Message ID", "Validation Status",
]

HEADER_FILL   = PatternFill("solid", start_color="4B3FA0")   # deep purple
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT     = Font(name="Arial", size=10)
ALT_FILL      = PatternFill("solid", start_color="F4F3FF")   # very light purple
WARN_FILL     = PatternFill("solid", start_color="FFF3CD")   # amber
ERR_FILL      = PatternFill("solid", start_color="F8D7DA")   # red
OK_FILL       = PatternFill("solid", start_color="D1FAE5")   # green
THIN_BORDER   = Border(
    left=Side(style="thin", color="D0CCEE"),
    right=Side(style="thin", color="D0CCEE"),
    top=Side(style="thin", color="D0CCEE"),
    bottom=Side(style="thin", color="D0CCEE"),
)

COL_WIDTHS = [5, 18, 16, 22, 16, 8, 18, 20, 36, 10, 12, 28, 18]


def _period_label(d: date) -> str:
    """Return a period string like '2025-Q1' or '2025-P1' depending on rotation."""
    if EXCEL_ROTATION_MONTHS == 3:
        q = (d.month - 1) // 3 + 1
        return f"{d.year}-Q{q}"
    period = (d.month - 1) // EXCEL_ROTATION_MONTHS + 1
    return f"{d.year}-P{period}"


def _workbook_path(d: date) -> Path:
    out_dir = Path(EXCEL_OUTPUT_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir / f"PF_Requests_{_period_label(d)}.xlsx"


def _sheet_name(d: date) -> str:
    return d.strftime("%Y-%m-%d")


def _get_or_create_workbook(path: Path) -> openpyxl.Workbook:
    if path.exists():
        return openpyxl.load_workbook(str(path))
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet
    logger.info("Created new workbook: {}", path)
    return wb


def _get_or_create_sheet(wb: openpyxl.Workbook, sheet_name: str):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    # Write header row
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_idx - 1]

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    logger.info("Created new sheet '{}' in workbook", sheet_name)
    return ws


def append_entry(entry: dict, validation_status: str = "Pending") -> tuple[str, int]:
    """
    Write one PF request entry to the correct workbook and daily sheet.

    Args:
        entry: dict with keys matching the agent's output schema.
        validation_status: 'OK', 'Warning', 'Error', or 'Pending'.

    Returns:
        (workbook_path_str, row_number)
    """
    today    = date.today()
    wb_path  = _workbook_path(today)
    wb       = _get_or_create_workbook(wb_path)
    ws       = _get_or_create_sheet(wb, _sheet_name(today))

    row = ws.max_row + 1
    serial = row - 1   # row 1 is header

    values = [
        serial,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        entry.get("employee_index", ""),
        entry.get("name", ""),
        entry.get("department", ""),
        entry.get("grade", ""),
        entry.get("pf_balance", ""),
        entry.get("amount"),
        entry.get("reason", ""),
        entry.get("source", "email"),
        entry.get("confidence", ""),
        entry.get("gmail_id", ""),
        validation_status,
    ]

    fill = _status_fill(validation_status)
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font      = BODY_FONT
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 9))

        if col_idx == 8 and value is not None:   # amount — right-align, number format
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")
        if col_idx == 7 and value is not None:   # pf balance
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")
        if col_idx == 13:   # validation status — coloured
            cell.fill = fill
            cell.font = Font(name="Arial", size=10, bold=True)

        # Alternating row fill for non-status columns
        if col_idx != 13 and row % 2 == 0:
            cell.fill = ALT_FILL

    ws.row_dimensions[row].height = 18
    wb.save(str(wb_path))
    logger.info("Entry #{} written to {} / sheet '{}'", serial, wb_path.name, _sheet_name(today))
    return str(wb_path), row


def update_validation_status(wb_path: str, row: int, status: str, notes: str = ""):
    """Update the validation status cell of an existing row."""
    today      = date.today()
    sheet_name = _sheet_name(today)

    wb = openpyxl.load_workbook(wb_path)

    if sheet_name not in wb.sheetnames:
        wb.close()
        return

    ws   = wb[sheet_name]
    cell = ws.cell(row=row, column=13)
    cell.value = f"{status}: {notes}" if notes else status
    cell.fill  = _status_fill(status)
    cell.font  = Font(name="Arial", size=10, bold=True)

    wb.save(wb_path)
    wb.close()
    
def _status_fill(status: str) -> PatternFill:
    s = (status or "").lower()
    if "ok" in s or "valid" in s or "pass" in s:
        return OK_FILL
    if "warn" in s or "flag" in s:
        return WARN_FILL
    if "error" in s or "fail" in s or "invalid" in s:
        return ERR_FILL
    return PatternFill()   # no fill for pending / unknown


def get_today_entries() -> list[dict]:
    """Read all entries from today's sheet for validation."""
    today   = date.today()
    path    = _workbook_path(today)
    sheet   = _sheet_name(today)

    if not path.exists():
        logger.warning("No Excel file found for today at {}", path)
        return []

    # Must open without read_only so we can later write back to same file
    wb = openpyxl.load_workbook(str(path), data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        logger.warning("No sheet '{}' found in {}", sheet, path.name)
        return []

    ws   = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    headers = [str(h).strip() if h else "" for h in rows[0]]
    entries = []
    for row_idx, row in enumerate(rows[1:], start=2):
        # Skip completely empty rows
        if all(v is None for v in row):
            continue
        record = dict(zip(headers, row))
        record["_row"]     = row_idx
        record["_wb_path"] = str(path)
        entries.append(record)

    logger.info("Found {} entries in today's sheet for validation", len(entries))
    return entries
